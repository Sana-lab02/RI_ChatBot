import pandas as pd
import json
from bot.Keywords import column_aliases, credential_terms, column_to_names, allowed_update_columns, scan_entry_triggers, exit_commands, inventory_triggers
from bot.bot_matchers import find_best_row, find_best_column, parse_requested_columns
from bot.bot_utils import *
from bot.scan_pred import ScanPredictor
from bot.scan_history import ScanHistory
from bot.flow_engine import FlowEngine
from bot.inventory import InventoryManager
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import sqlite3
import inspect
from openpyxl import load_workbook
import os
from datetime import datetime

#Confidence Threasholds
RETAILER_HIGH = 70
RETAILER_MEDIUM = 40
COLUMN_HIGH = 0.75
COLUMN_MEDIUM = 0.60
TROUBLE_THREASHOLD = 0.40

MAX_INFO_TURNS = 3
class RetailBot:
    def __init__(self, db_path="retailers.db"):
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        cursor = self.conn.cursor()
        
        self.df_customer_info = pd.read_sql_query("SELECT * FROM retailers", self.conn)
        self.df_trouble = pd.read_sql_query("SELECT * FROM troubleshooting", self.conn)
        
        self.awaiting_info = None
        self.awaiting_info_turns = 0

        self.awaiting_shipping = None
        self.awaiting_shipping_turns = 0

        self.pending_retailer = None
        self.pending_column = None
        self.awaiting_multi_info = None

        self.awaiting_confirmation = None
        self.awaiting_confirmation_turns = 0
        self.last_user_input = None
        self.awaiting_retailer = False
        self.exit_commands = ["quit", "exit", "bye"]

        self.awaiting_parcel = None

        self.pending_action = None
        self.new_ipad = None
        self.new_sensor = None
        self.awating_equipment_choice = None
        self.awaiting_manual_enter = None

        self.troubleshooting_flows = self.load_flows()
        self.active_troubleshooting = None
        self.awaiting_flow_retailer = False
        self.pending_flow_id = None

        self.predictor = ScanPredictor(self.conn)
        self.scan_history = ScanHistory(self.conn)

        self.active_scan_entry = False

        self.inventory = InventoryManager(self.conn)
        


        self.df_trouble['clean_question'] = self.df_trouble['question'].astype(str).apply(clean_text_tfidf)
        self.vectorizer_trouble = TfidfVectorizer(ngram_range=(1, 2))
        self.tfidf_trouble = self.vectorizer_trouble.fit_transform(self.df_trouble['clean_question'])
        self.column_names = list(column_aliases.keys())
        column_docs = [" ".join([col] + column_aliases.get(col, [])) for col in self.column_names]
        self.intent_vectorizer = TfidfVectorizer(ngram_range=(1, 2))
        if column_docs:
            self.tfidf_columns = self.intent_vectorizer.fit_transform([clean_text(doc) for doc in column_docs])
        else:
            self.tfidf_columns = None


    def process_input(self, user_input, role="user"):
        """Handles yes/no confirmation in Flask"""
        user_input = str(user_input).strip()

#------ Handle forms for inventory management ------------------------------------------------------------------------        
        if user_input.startswith("open_form "):
            form_id = user_input.split(" ", 1)[1].strip()

            if form_id == "inventory_add_device":
                return {"reply": self.inventory.add_device_form()}
            if form_id == "inventory_remove_device":
                return {"reply": self.inventory.remove_device_form()}
            if form_id == "inventory_checkout":
                return {"reply": self.inventory.checkout_form()}
            if form_id == "inventory_checkin":
                return {"reply": self.inventory.checkin_form()}

        if is_inventory_request(user_input):
            return {"reply": self.inventory.dashboard_form(is_admin=(role == "admin"))}
        
        if "update retailer" in user_input.lower():
            return {"reply": self.update_retailer_form()}

#------- Confirmations and Pending States ---------------------------------------------------------------------------------       
        scan_form_response = self.handle_scan_entry_input(user_input)
        if scan_form_response:
            return scan_form_response

        if self.active_scan_entry:
            return self.handle_scan_entry_mode(user_input)
        
        flow_resume = self.resume_flow_with_retailer(user_input)
        if flow_resume:
            return flow_resume
        
        flow_response = self.handle_flows(user_input)
        if flow_response:
            return flow_response
        
        if self.is_known_troubleshooting_request(user_input):
            return self.list_known_troubleshooting()
        
        if is_retailer_info_question(user_input):
            return self.handle_retailer_input(user_input)
        
        trouble_answer = self.get_troubleshooting_answer(user_input)
        if trouble_answer:
            return trouble_answer
        
        if user_input in ["help", "h", "?"]:
            return self.handel_help()
        
        if is_retailer_info_question(user_input):
            return self.handle_retailer_input(user_input)

        if self.awaiting_confirmation:
            self.awaiting_confirmation_turns += 1
            return self.handle_confirmation(user_input)
        
        if self.awaiting_multi_info:
            return self.get_mutliple_info(user_input)
                
        if self.awaiting_shipping:
            self.awaiting_shipping_turns += 1
            return self.handle_shipping_input(user_input)
        
        if self.awaiting_retailer:
            return self.handle_retailer_input(user_input)
        
#------ Parcel Shipper Flow -------------------------------------------        
        if self.awaiting_parcel:
            return self.handle_parcel_flow(user_input)
        
        if is_parcel_shipper_request(user_input):
            row_index, retailer, score = find_best_row(
                user_input, self.df_customer_info, threshold=60
            )

            if not retailer:
                return "Which retailer is this shipment for?"
            
            self.awaiting_parcel = {
                "retailer": retailer,
                "use_equipment_on_file": None,
                "manual_items": None,
                "shipping_method": None
            }

            return "Do you want to use the equipment on file for this retailer?"
        
#------- Update Intent -----------------------------------------------------
        
        if detect_multiple_updates(user_input) or is_note_addition(user_input):
            return self.handle_multi_update(user_input, author="Bot")
        
#------- Equipment -----------------------------------------------------------
        
        if self.pending_action == "new_equipment" and self.new_ipad is None:
            self.new_ipad = user_input
            return "What is the new sensor serial?"

        if self.pending_action == "new_equipment" and self.new_sensor is None:
            self.new_sensor = user_input

            row_index, retailer, score = find_best_row(self.last_user_input, self.df_customer_info)
            self.send_new_equipment(retailer, self.new_ipad, self.new_sensor)

            self.pending_action = None
            self.new_ipad = None
            self.new_sensor = None

            return f"New equipment saved for {retailer} and old equipment moved to returning."

        if any(x in user_input.lower() for x in ["new equipment", "update equipment", "replace equipment", "send new equipment"]):
            self.pending_action = "new_equipment"
            self.last_user_input = user_input
            self.new_ipad = None
            self.new_sensor = None
            return "What is the new iPad number?"
        
#------- Scans and Predictions ----------------------------------------------------------------------
        result = self.route_scan_request(user_input)
        if result is not None:
            return result

        return self.answer(user_input)
    

    def answer(self, user_input):
        self.last_user_input = user_input

        if is_retailer_info_question(user_input):
            return self.get_mutliple_info(user_input)

        if not is_retailer_info_question(user_input):
            return self.get_troubleshooting_answer(user_input)
        

        ranked = find_best_row(user_input, self.df_customer_info, threshold=40)
        row_index, retailer_name, r_score = ranked

        print("Debug... ", row_index, retailer_name, r_score)
        if retailer_name is None:
            if is_retailer_info_question(user_input):
                return f"Sorry I couldn't find that retailer. Can you double-check the name?"
            return self.get_troubleshooting_answer(user_input)
        
        # Medium confidence -> ask user to confirm
        if r_score < RETAILER_MEDIUM:
            self.awaiting_confirmation = {
                "row_index": row_index,
                "retailer_name": retailer_name
            }
            return f"I think you mean {retailer_name}. Is that correct? (yes/no)"
        
        if retailer_name is None:
            return self.get_troubleshooting_answer(user_input)
        

        # High confidence -> auto-select retailer and look for column
        result = find_best_column(user_input, column_aliases)

        if not result:
            self.awaiting_info = {
                "row_index" : row_index,
                "retailer_name": retailer_name
            }
            self.awaiting_info_turns = 0
            return f"I foud {retailer_name}. What information do you need?"

        col_name, col_score = result

        if col_score < COLUMN_MEDIUM:
            return f"For {retailer_name}, did you want {col_name.lower()}?"
    
        return self.get_column_value(row_index, col_name, retailer_name)
    
    def get_column_value(self, row_index, col_name, retailer_name):
        # Auto-fetch value if high enough confidence
        actual_cols = {c.lower().strip(): c for c in self.df_customer_info.columns}
        real_col = actual_cols.get(col_name.lower().strip(), col_name)

        if not real_col:
            return f"Sorry, I could't find that information for {retailer_name}."
        

        value = self.df_customer_info.at[row_index, real_col]

        if pd.isna(value) or str(value).strip() == "":
            return f"No information stored for {retailer_name}."

        self.reset_state()
        friend_col = column_to_names.get(
            real_col,
            real_col.replace("_", " ").title()
        )
        return f"{friend_col} for {retailer_name} is: {value}"


    def get_troubleshooting_answer(self, user_input):
        """Use TF-IDF to find the closest troubleshooting answer from df_trouble"""
        if is_retailer_info_question(user_input):
            return None
        try:
            clean_input = clean_text(user_input)
            user_vec = self.vectorizer_trouble.transform([clean_input])
            similarities = cosine_similarity(user_vec, self.tfidf_trouble).flatten()
            best_index = similarities.argmax()
            score = similarities[best_index]

            if score < TROUBLE_THREASHOLD:
                return None

            return self.df_trouble['question'].iloc[best_index] + ": " + self.df_trouble['answer'].iloc[best_index]
        except Exception as e:
            print("Debug...", e)
            return "Sorry, something went wrong while searching for a troubleshooting answer."
    


    def handle_confirmation(self, user_input):
        user_input = clean_text(user_input)
        data = self.awaiting_confirmation

        if user_input in ["yes", "y"]:
            self.awaiting_confirmation = None
            
            self.awaiting_info = {
                "row_index": data["row_index"],
                "retailer_name": data["retailer_name"]
            }
            self.awaiting_info_turns = 0

            return f"Great. What information do you need for {data['retailer_name']}"
        
        elif user_input in ["no", "n"]:
            self.awaiting_confirmation = None
            self.awaiting_retailer = True
            return "Okay, please tell me the correct retailer name"
        else:
            return "Please answer yes or no."
        
    def answer_with_locked_retailer(self, user_input, row_index, retailer_name):
        result = find_best_column(user_input, column_aliases)

        if not result:
            return f"I found {retailer_name}, but what information do you need?"
        
        col_name, col_score = result

        if col_score < COLUMN_MEDIUM:
            return f"For {retailer_name}, did you want {col_name.lower()}?"
        
        return self.get_column_value(row_index, col_name, retailer_name)
    

    def handle_info_request(self, user_input):
        data = self.awaiting_info
        row_index = data["row_index"]
        retailer_name = data["retailer_name"]

        result = find_best_column(user_input, column_aliases)

        if not result :
            return "I still couldn't tell what info you need. Try saying password, username, or account number"
        
        col_name, col_score = result

        if col_score < COLUMN_MEDIUM:
            return f"For {retailer_name}, did you want {col_name}?"
        
        self.awaiting_info = None
        return self.get_column_value(row_index, col_name, retailer_name)

    def reset_state(self):
        self.awaiting_info = None
        self.awaiting_info_turns = 0
        self.awaiting_confirmation = None
        self.awaiting_confirmation_turns = 0

    def handle_retailer_input(self, user_input):
        self.awaiting_retailer = False

        ranked = find_best_row(user_input, self.df_customer_info, threshold=50)
        row_index, retailer_name, score = ranked

        if retailer_name is None:
            self.awaiting_confirmation = True
            return "Sorry, I still couldn't find that retailer. Please try again"
        
        requested_cols = parse_requested_columns(user_input, column_aliases)
        
        if len(requested_cols) > 1:
            self.awaiting_multi_info = {
                "retailer": retailer_name
            }
            return self.get_mutliple_info(user_input)
        
        if is_retailer_info_question(user_input):
            print(">>> FULL INFO PATH for retailer:", retailer_name)

            row = self.df_customer_info.iloc[row_index]
            requested_field = extract_requested_field(user_input)

            if requested_field and requested_field in row.index:
                label = requested_field.replace("_", " ").title()
                value = row[requested_field]

                if pd.notna(value) and str(value).strip():
                    return f"{retailer_name}'s {label} is {value}."
                else:
                    return f"I dont have a {label} on file for {retailer_name}."

            lines = self.format_retailer_row(row.to_dict())
            return f"All info for {retailer_name}:\n" + "\n".join(lines)
        
        self.awaiting_info = {
            "row_index": row_index,
            "retailer_name": retailer_name
        }
        self.awaiting_info_turns = 0

        return f"Got it {retailer_name}. What information do you need?"
    
    def parse_shipping_method(self, user_input):
        shipping = user_input.strip().lower()

        if shipping in ["ground", "g"]:
            return "Ground"
        if shipping in ["2 day", "two day", "2d", "two d"]:
            return "2 Day"
        if shipping in ["Overnight"]:
            return "Overnight"
        
        return None
    
    def handle_parcel_flow(self, user_input):
        state = self.awaiting_parcel
        if not state:
            return "No parcel operation in progress."
        
        text = user_input.lower()

        if state["use_equipment_on_file"] is None:
            if text in ["yes", "y"]:
                state["use_equipment_on_file"] = True
                return "What shipping method?"
            elif text in ["no", "n"]:
                state["use_equipment_on_file"] = False
                return "What items are being shipped?"
            else:
                return "Please answer yes or no."
            
        if state["use_equipment_on_file"] is False and state["manual_items"] is None:
            state["manual_items"] = user_input
            return "What shipping method?"
        
        if state["shipping_method"] is None:
            method = self.parse_shipping_method(user_input)
            if not method:
                return "Please choose Ground, 2 Day or Overnight"
            state["shipping_method"] = method
            result = self.handle_parcel_shipper(
                state["retailer"],
                shipping_method=method
            )

            self.awaiting_parcel = None
            return result

    def handle_parcel_shipper(self, user_input, shipping_method="Ground"):
        import re
        from openpyxl.drawing.image import Image


        ranked = find_best_row(user_input, self.df_customer_info, threshold=60)
        row_index, retailer_name, score = ranked

        if retailer_name is None:
            return "I couldn't find that retailer. please try agian"
        
        template_path = "/Users/phood/Template/Parcel_Shipper_Template.xlsx"
        if not os.path.exists(template_path):
            return "Parcel shipper file not found."
        
        wb = load_workbook(template_path)
        ws = wb.active

        log_path = "/Users/phood/Template/Logo/Picture1.png"

        if os.path.exists(log_path) and not ws._images:
            img = Image(log_path)
            img.anchor = "F3"
            ws.add_image(img)


        row = self.df_customer_info.iloc[row_index]

        #==== map the cells ===
        ws["B9"] = datetime.now().strftime("%m/%d/%Y")
        ws["G9"] = row['retailer']
        ws["G10"] = ws["G10"].value + f" {row['fitter']}"
        ws["G11"] = row['street']
        
        city = row['city']
        state = row['state']
        zip_code = str(int(float(row['zip_code']))) if pd.notna(row['zip_code']) else ""

        ws["G12"] = f"{city}, {state} {zip_code}"
        ws["G13"] = row['country']
        ws["B19"] = ws["B19"].value + f" {row['account_number']}"

        parcel_state = self.awaiting_parcel or {}

        if parcel_state.get("use_equipment_on_file") is False:
            shipment_text = parcel_state.get("manual_items", "Manual shipment")
        else:
            shipment_text = f"Trupad {row['ipad_number']} and {row['sensor_serial']}"

        ws["B23"] = f"{ws['B23'].value} {shipment_text}"
        ws["C30"] = "X"
        ws["B32"] = shipping_method
        ws["B33"] = "NO"

        retailer_safe = re.sub(r'[^a-zA-z0-9_-]', '_', retailer_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        output_dir = os.path.join(os.getcwd(), "generated")
        os.makedirs(output_dir, exist_ok=True)

        output_path = os.path.join(
            output_dir,
            f"parcel_{retailer_safe}_{timestamp}.xlsx"
        )

        wb.save(output_path)

        return {
            "file": output_path,
            "filename": os.path.basename(output_path)
        }
    
    def update_customer_info(self, retailer_name, updates: dict):
        retailer_name = str(retailer_name).strip()

        row = self.df_customer_info[self.df_customer_info['retailer'] == retailer_name]
        if row.empty:
            return f"Retailer '{retailer_name}' not found."
        
        for col in val in updates.items():
            if col in self.df_customer_info.columns:
                self.df_customer_info.loc[self.df_customer_info['retailer'] == retailer_name, col] = val

        conn = self.conn
        cursor = conn.cursor()
        for col, val in updates.items():
            if col in self.df_customer_info.columns:
                cursor.execute(f"UPDATE retailers set {col} = ? WHERE retailer = ?", (val, retailer_name))
        conn.commit()

        return f"Updated {retailer_name} successfully."
    
    def lookup_retaier_info(self, user_input):
        requested_cols = parse_requested_columns(user_input, column_aliases)

        row_index, retailer, score = find_best_row(user_input, self.df_customer_info)
        if retailer is None:
            return "Sorry I could't find that retailer."
        
        
        row = self.df_customer_info.iloc[row_index]
        results = []
        for col in requested_cols:
            val = row.get(col, "")
            if pd.isna(val) or val == "":
                val = "Not on file"
            results.append(f"{col.replace('_',' ').title()}: {val}")

        return f"{retailer} info: " + "; ".join(results)
    
    def get_mutliple_info(self, user_input):
        try:

            if not self.awaiting_multi_info:
                _, retailer, score = find_best_row(user_input, self.df_customer_info, threshold=70)

                if not retailer:
                    return "I couldn't find retailer."
            
                self.awaiting_multi_info = {
                    "retailer": retailer
                }

            retailer = self.awaiting_multi_info["retailer"]

            request_cols = parse_requested_columns(user_input, column_aliases)

            if not request_cols:
                return "What information do you want about them?"
                
            row_index, _, score = find_best_row(retailer, self.df_customer_info, threshold=80)

            if row_index is None:
                self.awaiting_multi_info = None
                return "I lost track of that retailer. Please ask again."
            
            row = self.df_customer_info.iloc[row_index]

            if str(row.get("retailer", " ")).strip() != retailer:
                self.awaiting_multi_info = None
                return "Internal error: retailer mismatch. Please ask again"
            

            print("===== DEBUG ROW CHECK =====")
            print("User input:", user_input)
            print("Expected retailer:", retailer)
            print("Row index:", row_index)
            print("Row retailer:", row.get("retailer"))
            print("===========================")
            responses = []
            for col in request_cols:
                value = row.get(col, "")
                if pd.isna(value) or str(value).strip() == "":
                    value = "Not on file"
                responses.append(f"{col.replace('_',' ').title()}: {value}")

            self.awaiting_multi_info = None

            return f" Here is the information for {retailer}:\n" + "\n".join(responses)
        finally:
            self.awaiting_multi_info = None
    
    def handle_multi_update(self, user_input, author="Bot"):
        row_index, retailer, score = find_best_row(user_input, self.df_customer_info)

        if not retailer:
            return "I couldnt find that retialer."
        
        updates = detect_multiple_updates(user_input)
        if not updates:
            return "I couldn't detect what you want me to update"
        
        retailer_db = retailer.strip().lower()
        cursor = self.conn.cursor()
        changed = []

        for column, value in updates.items():
            if column not in allowed_update_columns:
                continue
            
            if column in ["notes", "jane_notes"]:
                cursor.execute(
                    f"SELECT {column} FROM retailers WHERE retailer = ?",
                    (retailer,)
                )
                row = cursor.fetchone()
                existing = row[0] if row and row[0] else ""
                value = append_note(existing, value, author)
            
            cursor.execute(
                f"UPDATE retailers SET {column} = ? WHERE retailer = ?",
                (value, retailer) 
            )

            if cursor.rowcount > 0:
                changed.append(column)

        self.conn.commit()

        if not changed:
            return f"No updates were applied for {retailer}"
        
        update_fields = ", ".join(col.replace("_", " ") for col in updates)
        return f"Updated {update_fields} for {retailer}."
    
    def route_scan_request(self, user_input):
        text = user_input.lower()

        if not any(k in text for k in ["scan", "scans", "history", "how many", "count", "predict", "forecast", "projection"]):
            return None

        row_index, retailer, score = find_best_row(text, self.df_customer_info, threshold=60)

        if any(k in text for k in ["predict", "forecast", "future", "projection"]):
            months = extract_months(text) or 3

            if not retailer:
                return f"Which retailer do you want to predict scans for?"
            
            if not self.predictor.retailer_exists(retailer):
                return f"Retailer {retailer} not found"
            
            result = self.predictor.predict_scans_with_graph(retailer, months)
            preds = result["predictions"]

            if preds["predicted_scan_count"].sum() == 0:
                return f"Not enough scan history for {retailer}"
            
            text_out = "\n".join(
                f"{row['ds'].strftime('%b %Y')}: {int(row['predicted_scan_count'])}"
                for _, row in preds.iterrows()
            )

            return {
                "text": f"Predicted scans for {retailer}:\n{text_out}",
                "image": result["image"]
            }
        
        if any(k in text for k in ["how many", "count", "total", "number", "past scan", "history"]):

            if not retailer:
                return "Which retailer?"
            
            if "last" in text or "months" in text:
                months = extract_months(user_input) or 3
                df = self.scan_history.scans_last_n_months(retailer, months)
                
            else:
                df = self.scan_history.scans_full_history(retailer)
            
            if df.empty:
                return f"No scan history found for {retailer}"
            
            retail_month = self.scan_history.scans_monthly_history(retailer)
            text_out = self.scan_history.format_monthly_counts(retail_month)
            graph_img = self.scan_history.plot_scan_history(df, retailer)
            
            return {
                "text": f"Scan history for {retailer}:\n{text_out}",
                "image": graph_img
            }
    
        return None
    
    def handel_help(self):
        help_commands = [
            "Customer Info lookup: 'What is the useranme and password for forever me? (Try all info)\n",
            "Check scan history: Scan history for 3 months for images boutique (or just scan history for)\n",
            "Predict future scans: Predict scans for 4 months for images boutique (up to 12 months)\n"
        ]

        help_text = "Here are some examples of how you can interact with me! I recommend copy and pasting:\n\n"
        help_text += "\n".join(f"- {ex}" for ex in help_commands)
        return help_text

    def refresh_customer_db(self):
        self.df_customer_info = pd.read_sql_query("SELECT * FROM retailers", self.conn)


    def load_flows(self, path="Troubleshooting_flows/Troubleshooting.json"):
        if not os.path.exists(path):
            print(f"Flow file not found: {path}")
            return {}
        
        with open(path, "r") as f:
            data = json.load(f)
        
        flows = {}
        for flow_id, flow_data in data.items():
            flows[flow_id] = flow_data
        return flows
    
    def handle_flows(self, user_input):
        text = clean_text(user_input.lower())

    # If a troubleshooting flow is already active, handle it
        if self.active_troubleshooting:
            response = self.active_troubleshooting.handle_input(user_input)
            if response:
                return response
            else:
                self.active_troubleshooting = None

        # Loop through flows to see if any triggers match
        for flow_id, flow_data in self.troubleshooting_flows.items():
            triggers = flow_data.get("triggers", [])
            for trig in triggers:
                trig_clean = clean_text(trig.lower())
                if trig_clean in text:
                    context = {}

                    row_index, retailer_name, score = find_best_row(user_input, self.df_customer_info)
                    if not retailer_name:
                        self.awaiting_flow_retailer = True
                        self.pending_flow_id = flow_id
                        return "Which retaielr are you trying to log into?"
                    
                    row = self.df_customer_info.iloc[row_index]
                    context = {
                        "retailer": retailer_name,
                        "ri_app_password": row.get("ri_app_password")
                    }

                    # Start the troubleshooting flow with context
                    self.active_troubleshooting = FlowEngine(
                        self.troubleshooting_flows, context=context
                    )
                    return self.active_troubleshooting.start_flow(flow_id)
        
        return None
    
    def resume_flow_with_retailer(self, user_input):
        if not self.awaiting_flow_retailer:
            return None
        
        row_index, retailer_name, score = find_best_row(
            user_input, self.df_customer_info, threshold=60
        )

        if not retailer_name:
            return "Sorry, I still couldn't find that retailer. Please try again."
        
        self.active_troubleshooting = FlowEngine(
            self.troubleshooting_flows,
            context={
                "retailer": retailer_name,
                "ri_app_password": self.df_customer_info.loc[row_index, "ri_app_password"]
            }
        )

        self.awaiting_flow_retailer = False

        return self.active_troubleshooting.start_flow(self.pending_flow_id)


    def list_known_troubleshooting(self):
        topics = set()

        if self.df_trouble.empty:
            return "I dont have any troubleshooting topics saved yet"
        
        seen = {}
        for q in self.df_trouble["question"].dropna().astype(str):
            normalized = clean_text(q)
            if normalized not in seen:
                seen[normalized] = q.strip()

        topics = sorted(seen.values(), key=str.lower)

        response = "Here are the troubleshooting issues I can help with;\n\n"
        response += "\n".join(f"- {t}" for t in topics)

        return response
    
    def is_known_troubleshooting_request(self, user_input):
        text = clean_text(user_input.lower())
        return any(trigger in text for trigger in trouble_shooting_triggers)
    
    def handle_scan_entry_input(self, user_input):
        text = user_input.lower()

        if any(t in text for t in scan_entry_triggers):
            self.active_scan_entry = True

            return { 
                "type": "scan_entry_form"
            }
        
        return None
    
    def handle_scan_entry_mode(self, user_input):
        if user_input.lower() in self.exit_commands:
            self.active_scan_entry = False
            return "Scan entry complete!"
        
        return "Scan added successfully."
    

    def add_scan_form(self):
        """
        Returns a unified form structure for adding scans.
        This matches the structure of update_retailer_form().
        """
        retailers = [str(r) for r in self.df_customer_info["retailer"].dropna().unique()]
        
        return {
            "type": "session form",
            "form_id": "add_scan",
            "title": "Add New Scan",
            "fields": [
                {
                    "name": "retailer",
                    "type": "text",
                    "label": "Retailer",
                    "placeholder": "Type to search...",
                    "options": retailers
                },
                {
                    "name": "date",
                    "type": "text",
                    "label": "Date",
                    "placeholder": "MM/DD/YYYY",
                    "options": []  # No autocomplete for date
                },
                {
                    "name": "count",
                    "type": "number",
                    "label": "Scan Count",
                    "placeholder": "1",
                    "options": []  # No autocomplete for count
                }
            ],
            "dynamic_fields": None,  # Not needed for add scan
            "value_field": None,  # Not needed for add scan
            "submit_label": "Add Scan",
            "submit_template": "add scan {0} {1} {2}",  # {0}=retailer, {1}=date, {2}=count
            "success_message": "✓ Scan added successfully",
            "buttons": [
                {"text": "Add Scan", "action": "submit"},
                {"text": "Cancel", "action": "exit"}
            ]
        }

    def handle_form_submission(self, payload, role="user"):
        form_id = payload.get("form_id")
        data = payload.get("data", {})
        
        if form_id == "update_retailer":
            return self.apply_retailer_updates(data)
        
        if form_id == "add_scan":
            return self.apply_scan_entry(data)
        
        if form_id == "add_retailer":
            return self.add_new_retailer(data)
        
        if form_id == "add_note":
            return self.add_new_note(data)
        
        if form_id and form_id.startswith("inventory_"):
            is_admin = (role == "admin")

            if form_id == "inventory_add_device":
                if not is_admin:
                    return {"text": "⚠ Admin privileges required"}
                result = self.inventory.add_device(data)
                result["reply"] = self.inventory.dashboard_form(is_admin=is_admin)
                return result
            
            if form_id == "inventory_remove_device":
                if not is_admin:
                    return {"text": "⚠ Admin privileges required"}
                result = self.inventory.retire_device(data)
                result["reply"] = self.inventory.dashboard_form(is_admin=is_admin)
                return result
            
            return self.inventory.handle_form_submission(payload, is_admin=is_admin)
        
        return {"text": "Unknown form submission."}


    def apply_scan_entry(self, data):
        """
        Process scan entry from the unified form.
        data = {
            "retailer": "Walmart",
            "date": "01/15/2024",
            "count": "5"
        }
        """
        retailer = data.get("retailer", "").strip()
        date = data.get("date", "").strip()
        count = data.get("count", "").strip()
        
        if not retailer or not date or not count:
            return {"text": "⚠ Please fill all fields"}
        
        try:
            # Parse the date
            from datetime import datetime
            date_obj = datetime.strptime(date, "%m/%d/%Y")
            
            # Validate count is a number
            scan_count = int(count)
            
            # Insert into scan history
            cursor = self.conn.cursor()
            cursor.execute(
                "INSERT INTO scan_history (retailer, scan_date, scan_count) VALUES (?, ?, ?)",
                (retailer, date_obj.strftime("%Y-%m-%d"), scan_count)
            )
            self.conn.commit()
            
            return {"text": f"✓ Added {scan_count} scan(s) for {retailer} on {date}"}
        
        except ValueError as e:
            return {"text": f"⚠ Invalid date format. Please use MM/DD/YYYY"}
        except Exception as e:
            print(f"Error adding scan: {e}")
            return {"text": "⚠ Failed to add scan"}
    
    def update_retailer_form(self):
        retailers = [str(r) for r in self.df_customer_info["retailer"].dropna().unique()]
        fields_to_update = sorted([str(f) for f in allowed_update_columns])
        return {
            "type": "session form",
            "form_id": "update_retailer",
            "title": "Update Retailer Information",
            "fields": [
                {
                    "name": "retailer",
                    "type": "dropdown",
                    "label": "Retailer",
                    "options": retailers
                }
            ],
            "dynamic_fields": {
                "label": "Fields to update",
                "options": fields_to_update
            },
            "buttons": [
                {"text": "Save", "action": "submit"},
                {"text": "Exit", "action": "exit"}
            ]
        }
    
    def apply_retailer_updates(self, data):
        retailer = (data.get("retailer") or "").strip()
        if not retailer:
            return {"text": "⚠ Missing retailer."}

        field = (data.get("field") or "").strip()
        value_present = "value" in data
        value = data.get("value")

        updates = {}

        if field and value_present:
            updates[field] = value
        else:
            # fallback if your frontend ever sends dict updates
            if isinstance(data.get("updates"), dict):
                updates = data["updates"]
            elif isinstance(data.get("updates"), list):
                for item in data["updates"]:
                    k = (item.get("field") or item.get("name") or "").strip()
                    if k:
                        updates[k] = item.get("value")

        print("ALLOWED:", allowed_update_columns)
        print("UPDATES:", updates)

        # Allow-list filter
        safe_updates = {k: v for k, v in updates.items() if k in allowed_update_columns}
        if not safe_updates:
            return {"text": "⚠ None of the submitted fields are allowed to be updated."}

        cur = self.conn.cursor()

        # confirm retailer exists
        cur.execute("SELECT 1 FROM retailers WHERE retailer = ? LIMIT 1", (retailer,))
        if not cur.fetchone():
            return {"text": f"⚠ Retailer '{retailer}' not found."}

        rows_changed = 0
        for col, val in safe_updates.items():
            cur.execute(f"UPDATE retailers SET {col} = ? WHERE retailer = ?", (val, retailer))
            rows_changed += cur.rowcount

        self.conn.commit()
        self.refresh_customer_db()

        if rows_changed == 0:
            return {"text": f"⚠ No rows updated for {retailer} (value may be unchanged)."}
        return {"text": f"✓ Updated {', '.join(safe_updates.keys())} for {retailer}."}

    
        
